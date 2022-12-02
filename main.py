import pandas as pd
import numpy as np
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# setting selenium options
options = Options()
options.headless = True # keep selenium from opening browser window
driver = webdriver.Firefox(executable_path='C:/Users/Noah/AppData/Local/Programs/Python/Python310/geckodriver.exe', options=options)

# import the list of urls as a dataframe
url_list = pd.read_excel('C:/users/1sled/desktop/state_list', sheet_name='Sheet1', header=0)

# import the workbook we'll be editing
ws = load_workbook('C:\users\1sled\desktop\state_list')
ws = ws['Sheet1']


# copying the column containing the number of bids that was scraped the last time this script was run
old_num_bids_list = np.array(ws['D2:D76'])
													
# create empty array to fill up with each website's current number of bids
num_bids_list = np.array([])
													 
# loop through each url
for ind in url_list.index:
	# if it's a CA state and a planetbids website
    if (url_list['STATE'][ind] == 'California') and ((url_list['WEBSITE'][ind]).split('/')[2] == 'www.planetbids.com'):
        
        # generate the full url to visit			   
        url = 'pbsystem.planetbids.com/portal/' + (url_list['WEBSITE'][ind])[-5:] + '/bo/bo-search'
        driver.get(url)
        assert 'planetbids' in driver.title

        # finding the value we want
        element = driver.find_element(By.CLASS_NAME, "bids-table-filter-message")
        element = element.text
		# should return 'Found # bids'

		# the current number of bids
        num_bids = element.split(' ')[1] # the number of bids
        num_bids_list = np.append(num_bids_list, num_bids)
    
    # if it's not a CA state or planetbids site, ignore it and go to the next row
    else:
        num_bids_list = np.append(num_bids_list, '')
																															 
# subtract the difference between the old and new number of bids to find how many were added
num_of_new_bids = np.subtract(num_bids_list, old_num_bids_list)

# add it to the spreadsheet
url_list['number of new bids'] = num_of_new_bids
url_list.to_excel('c:/users/1sled/desktop/updated_states_list.xlsx', index=False)																	 
																			 
 
